 # imports
import os
import shutil
from datetime import datetime, timedelta
from PIL import Image
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import ipywidgets as widgets
from IPython.display import display

output = widgets.Output()

initial_source = 'd:/photos/2024/temp'
initial_target = 'd:/photos/2024/temp — renamed'
initial_offset = 0


def custom_print(*args, **kwargs):
    with output:
        print(*args, **kwargs)
        

def read_exif_data(file_path):
    try:
        img = Image.open(file_path)
        if hasattr(img, '_getexif') and img._getexif() is not None:
            exif_data = img._getexif()
            if 36867 in exif_data:
                original_date = datetime.strptime(exif_data[36867], '%Y:%m:%d %H:%M:%S')
                return original_date
            else:
                custom_print(f"Warning: No 'DateTimeOriginal' tag found in EXIF data for {file_path}.")
                return None
        else:
            custom_print(f"Warning: No EXIF data found in {file_path}.")
            return None
    except IOError as e:
        custom_print(f"Error: Could not open {file_path} - {e}")
        raise IOError(f"Could not open {file_path}: {e}")
    except Exception as e:
        custom_print(f"Error: An unexpected error occurred while reading {file_path} - {e}")
        return None


def get_new_filename(original_date, conflict_count, method, original_extension):
    if method == 'increment_seconds':
        new_date = original_date + timedelta(seconds=conflict_count)
        return f"{new_date.strftime('%Y.%m.%d - %H.%M.%S')}{original_extension}"
    elif method == 'add_counter':
        if conflict_count == 0:
            return f"{original_date.strftime('%Y.%m.%d - %H.%M.%S')}{original_extension}"
        else:
            return f"{original_date.strftime('%Y.%m.%d - %H.%M.%S')}_{conflict_count:03}{original_extension}"


def save_log_to_excel(log_entries, source_folder, conflict_resolution_method):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = f"rename_log_{conflict_resolution_method}_{timestamp}.xlsx"
    log_path = os.path.join(os.path.dirname(source_folder), log_filename)
    df_log = pd.DataFrame(log_entries, columns=['Original Name', 'New Name', 'Error', 'Conflict', 'Extension'])
    
    df_log.to_excel(log_path, index=False, sheet_name="Log")

    wb = openpyxl.load_workbook(log_path)
    ws = wb.active

    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[2].value == 'TRUE':
            for cell in row:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif row[3].value == 'TRUE':
            for cell in row:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(log_path)
    custom_print(f"Log saved to: {log_path}")
    return log_path


def generate_unique_filename(original_date, original_extension, conflict_resolution_method, 
                             name_mapping, destination_folder, is_dry_run):
    conflict_count = 0
    while True:
        new_filename = get_new_filename(original_date, conflict_count, conflict_resolution_method, original_extension)
        
        if is_dry_run:
            if new_filename not in name_mapping.values():
                break
        else:
            new_file_path = os.path.join(destination_folder, new_filename)
            if not os.path.exists(new_file_path):
                break
        
        conflict_count += 1
    
    if is_dry_run:
        name_mapping[original_date] = new_filename
    
    return new_filename, conflict_count

def generate_unique_movie_filename(original_date, original_extension, conflict_resolution_method, 
                             name_mapping, destination_folder, is_dry_run):
    conflict_count = 0
    while True:
        new_filename = get_new_filename(original_date, conflict_count, conflict_resolution_method, original_extension)
        
        if is_dry_run:
            if new_filename not in name_mapping.values():
                break
        else:
            new_file_path = os.path.join(destination_folder, new_filename)
            if not os.path.exists(new_file_path):
                break
        
        conflict_count += 1
    
    if is_dry_run:
        name_mapping[original_date] = new_filename
    
    return new_filename, conflict_count

    
def perform_file_operation(file_path, destination_folder, new_filename, rename_in_place):
    new_file_path = os.path.join(destination_folder if not rename_in_place else os.path.dirname(file_path), new_filename)
    
    if rename_in_place:
        os.rename(file_path, new_file_path)
    else:
        shutil.copy2(file_path, new_file_path)

def process_single_image(filename, source_folder, destination_folder, conflict_resolution_method, 
                         is_dry_run, rename_in_place, name_mapping, log_entries, conflict_files):
    file_path = os.path.join(source_folder, filename)
    original_extension = os.path.splitext(filename)[1]
    original_date = read_exif_data(file_path)
    
    if original_date is None:
        log_entries.append([filename, '', 'TRUE', 'FALSE', original_extension])
        return
    
    new_filename, conflict_count = generate_unique_filename(
        original_date, original_extension, conflict_resolution_method, name_mapping, destination_folder, is_dry_run
    )
    
    if conflict_count > 0:
        conflict_files.append((filename, new_filename))

    if not is_dry_run:
        perform_file_operation(file_path, destination_folder, new_filename, rename_in_place)
    
    log_entries.append([filename, new_filename, 'FALSE', 'TRUE' if conflict_count > 0 else 'FALSE', original_extension])

def get_movie_date(filepath, offset_hours):
    original_date = os.path.getmtime(filepath)
    date = datetime.fromtimestamp(original_date)
    new_date = date + timedelta(hours=offset_hours)
    return new_date

def process_single_movie(filename, source_folder, destination_folder, conflict_resolution_method, 
                         is_dry_run, rename_in_place, name_mapping, log_entries, conflict_files, offset_hours):
    file_path = os.path.join(source_folder, filename)
    original_extension = os.path.splitext(filename)[1]
    original_date = get_movie_date(file_path, offset_hours)
    
    new_filename, conflict_count = generate_unique_movie_filename(
        original_date, original_extension, conflict_resolution_method, name_mapping, destination_folder, is_dry_run
    )
    
    if conflict_count > 0:
        conflict_files.append((filename, new_filename))

    if not is_dry_run:
        perform_file_operation(file_path, destination_folder, new_filename, rename_in_place)
    
    log_entries.append([filename, new_filename, 'FALSE', 'TRUE' if conflict_count > 0 else 'FALSE', original_extension])

def print_image_summary(processed_files, exif_failures, conflict_files):
    conflicts = len(conflict_files)
    custom_print(f"Processed {processed_files} images, {conflicts} conflicts, {exif_failures} EXIF read failures.")
    
    if conflict_files:
        custom_print("Conflicts occurred in the following files:")
        for file in conflict_files:
            custom_print(f"{file}")


def print_movie_summary(processed_files, conflict_files):
    conflicts = len(conflict_files)
    custom_print(f"Processed {processed_files} movies, {conflicts} conflicts.")
    
    if conflict_files:
        custom_print("Conflicts occurred in the following files:")
        for file in conflict_files:
            custom_print(f"{file}")
            

def process_images_in_folder(source_folder, destination_folder, conflict_resolution_method, is_dry_run, rename_in_place, log_entries):
    if not is_dry_run:
        custom_print(f"\nRunning images")
    
    exif_failures = 0
    conflict_files = []
    name_mapping = {}

    if not is_dry_run and not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    image_files = [f for f in os.listdir(source_folder) if f.lower().endswith(('.jpg', '.jpeg'))]
    image_files.sort(key=lambda f: os.path.getmtime(os.path.join(source_folder, f)))
    processed_files = len(image_files)
    for filename in image_files:
        process_single_image(
            filename, source_folder, destination_folder, conflict_resolution_method,
            is_dry_run, rename_in_place, name_mapping, log_entries, conflict_files
        )

    print_image_summary(processed_files, exif_failures, conflict_files)

def process_movies_in_folder(source_folder, destination_folder, conflict_resolution_method, is_dry_run, rename_in_place, offset_hours, log_entries):
    if not is_dry_run:
        custom_print(f"\nRunning movies")
        
    conflict_files = []
    name_mapping = {}

    if not is_dry_run and not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    movie_files = [f for f in os.listdir(source_folder) if f.lower().endswith(('.mp4', '.avi'))]
    movie_files.sort(key=lambda f: os.path.getmtime(os.path.join(source_folder, f)))
    processed_files = len(movie_files)
    for filename in movie_files:
        process_single_movie(
            filename, source_folder, destination_folder, conflict_resolution_method,
            is_dry_run, rename_in_place, name_mapping, log_entries, conflict_files, offset_hours
        )

    print_movie_summary(processed_files, conflict_files)

def run(source_folder, destination_folder, conflict_resolution_method, is_dry_run, rename_in_place, offset_hours):
    log_entries = []
    output.clear_output()
    if is_dry_run:
        custom_print("DRY RUN\n\nTesting images counter")
        process_images_in_folder(source_folder, destination_folder, 'add_counter', is_dry_run, rename_in_place, log_entries)
        
        log_entries = []
        custom_print("\nTesting images adding seconds")
        process_images_in_folder(source_folder, destination_folder, 'increment_seconds', is_dry_run, rename_in_place, log_entries)
        
        log_entries = []
        custom_print("\n\nTesting movies counter")
        process_movies_in_folder(source_folder, destination_folder, 'add_counter', is_dry_run, rename_in_place, offset_hours, log_entries)
        
        log_entries = []
        custom_print("\nTesting movies adding seconds")
        process_movies_in_folder(source_folder, destination_folder, 'increment_seconds', is_dry_run, rename_in_place, offset_hours, log_entries)
    else:
        custom_print(f"Running with configuration: {locals()}")
        process_images_in_folder(source_folder, destination_folder, conflict_resolution_method, is_dry_run, rename_in_place, log_entries)
        process_movies_in_folder(source_folder, destination_folder, conflict_resolution_method, is_dry_run, rename_in_place, offset_hours, log_entries)
        save_log_to_excel(log_entries, source_folder, conflict_resolution_method)


def setup_widgets():
    layout = widgets.Layout(width='500px')

    source_folder_widget = widgets.Text(
        value=initial_source,
        description='Source:',
        disabled=False,
        layout=layout,
        style={'description_width': 'initial'}
    )

    destination_folder_widget = widgets.Text(
        value=initial_target,
        description='Target: ',
        disabled=False,
        layout=layout,
        style={'description_width': 'initial'}
    )

    offset_hours_widget = widgets.IntText(
        value=initial_offset,
        description='Offset: ',
        disabled=False,
        layout=layout,
        style={'description_width': 'initial'}
    )

    conflict_resolution_method = widgets.ToggleButtons(
        options=['add_counter', 'increment_seconds'],
        description='Conflict Resolution:',
        disabled=False,
        button_style='',
        tooltips=['Add a counter', 'Increment by seconds']
    )

    is_dry_run = widgets.Checkbox(
        value=True,
        description='Dry Run',
        disabled=False
    )

    rename_in_place = widgets.Checkbox(
        value=False,
        description='Rename In Place',
        disabled=False
    )

    run_button = widgets.Button(
        description="Dry Run",
        button_style='primary',
        tooltip='Simulate the process and detect conflicts',
        icon='check'
    )


    def update_button_appearance(change):
        if change['new']:
            run_button.icon = 'check'
            run_button.button_style = 'primary'
            run_button.description = "Dry Run"
            run_button.tooltip = 'Simulate the process and detect conflicts'
        else:
            run_button.icon = 'play'
            run_button.button_style = 'success'
            run_button.description = "Process Files"
            run_button.tooltip = 'Run the process'

    is_dry_run.observe(update_button_appearance, names='value')

    def on_run_button_clicked(b):
        run(
            source_folder=source_folder_widget.value,
            destination_folder=destination_folder_widget.value,
            conflict_resolution_method=conflict_resolution_method.value,
            is_dry_run=is_dry_run.value,
            rename_in_place=rename_in_place.value,
            offset_hours=offset_hours_widget.value
        )

    run_button.on_click(on_run_button_clicked)

    display(source_folder_widget, destination_folder_widget, offset_hours_widget, conflict_resolution_method, is_dry_run, rename_in_place, run_button, output)
